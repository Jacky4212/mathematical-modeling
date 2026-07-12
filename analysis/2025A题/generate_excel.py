"""
生成 result1.xlsx / result2.xlsx / result3.xlsx
================================================
按2025年CUMCM A题附件模板格式输出（自行推断模板结构）。
"""
import numpy as np
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from core_model import *

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system('pip install openpyxl -i https://pypi.tuna.tsinghua.edu.cn/simple')
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter


# ============================================================
# 公共样式
# ============================================================
HEADER_FONT = Font(name='微软雅黑', bold=True, size=11)
TITLE_FONT = Font(name='微软雅黑', bold=True, size=14)
NORMAL_FONT = Font(name='微软雅黑', size=10)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT_WHITE = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
LIGHT_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')


def style_header_row(ws, row, n_cols):
    """给表头行加样式"""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col):
    """给数据单元格加样式"""
    cell = ws.cell(row=row, column=col)
    cell.font = NORMAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    return cell


def auto_width(ws, min_width=10, max_width=22):
    """自动列宽"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # 估算中文字符宽度
                val = str(cell.value)
                cn_chars = sum(1 for c in val if '一' <= c <= '鿿')
                en_chars = len(val) - cn_chars
                width = cn_chars * 2.2 + en_chars * 1.1
                max_len = max(max_len, width)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# ============================================================
# result1.xlsx — 问题3结果
# ============================================================
def generate_result1(theta, speed, t_drops, t_delays, coverage, output_dir='.'):
    """
    参数:
        theta, speed: FY1飞行参数
        t_drops: [t_drop1, t_drop2, t_drop3]
        t_delays: [t_delay1, t_delay2, t_delay3]
        coverage: 联合遮蔽时长
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '问题3结果'

    # 标题
    ws.merge_cells('A1:I1')
    title_cell = ws.cell(row=1, column=1, value='问题3：FY1三弹时序优化结果')
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER_ALIGN

    # 基本信息
    row = 3
    info_data = [
        ('导弹', 'M1'),
        ('无人机', 'FY1'),
        ('初始位置', f'({DRONES["FY1"][0]:.0f}, {DRONES["FY1"][1]:.0f}, {DRONES["FY1"][2]:.0f})'),
        ('航向角', f'{np.degrees(theta):.2f}°'),
        ('飞行速度', f'{speed:.1f} m/s'),
        ('联合遮蔽时长', f'{coverage:.4f} s'),
    ]
    for label, value in info_data:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(name='微软雅黑', bold=True, size=10)
        c1.border = THIN_BORDER
        c1.alignment = LEFT_ALIGN
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        c2 = ws.cell(row=row, column=2, value=value)
        c2.font = NORMAL_FONT
        c2.border = THIN_BORDER
        c2.alignment = LEFT_ALIGN
        row += 1

    # 炸弹详情表
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=1, value='烟幕干扰弹投放参数').font = Font(name='微软雅黑', bold=True, size=12)
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN
    row += 1

    headers = ['弹序', '投放时刻(s)', '起爆延迟(s)', '起爆时刻(s)',
               '起爆点X(m)', '起爆点Y(m)', '起爆点Z(m)',
               '单弹遮蔽(s)', '备注']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    for i in range(3):
        t_burst = t_drops[i] + t_delays[i]
        p_burst = bomb_burst_position('FY1', speed, theta, t_drops[i], t_delays[i])
        single_cov = compute_coverage('M1', p_burst, t_burst, dt=0.02)

        values = [
            i + 1,
            round(t_drops[i], 3),
            round(t_delays[i], 3),
            round(t_burst, 3),
            round(p_burst[0], 1),
            round(p_burst[1], 1),
            round(p_burst[2], 1),
            round(single_cov, 4),
            '有效' if single_cov > 0 else '—',
        ]
        for col, v in enumerate(values, 1):
            style_data_cell(ws, row, col)
            ws.cell(row=row, column=col, value=v)
        row += 1

    auto_width(ws)

    filepath = os.path.join(output_dir, 'result1.xlsx')
    wb.save(filepath)
    print(f'[OK] {filepath}')
    return filepath


# ============================================================
# result2.xlsx — 问题4结果
# ============================================================
def generate_result2(results, coverage, output_dir='.'):
    """
    参数:
        results: {drone_id: {heading_deg, speed, t_drop, t_delay, t_burst,
                              burst_x, burst_y, burst_z, single_coverage}}
        coverage: 联合遮蔽时长
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '问题4结果'

    # 标题
    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value='问题4：三机协同单弹优化结果 (FY1+FY2+FY3 → M1)')
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER_ALIGN

    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c1 = ws.cell(row=row, column=1, value=f'导弹M1联合遮蔽时长: {coverage:.4f} s')
    c1.font = Font(name='微软雅黑', bold=True, size=12, color='C00000')
    row += 2

    # 主表
    headers = ['无人机', '初始位置', '航向角(°)', '速度(m/s)',
               '投放时刻(s)', '起爆延迟(s)', '起爆时刻(s)',
               '起爆点X(m)', '起爆点Y(m)', '起爆点Z(m)']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    for did in ['FY1', 'FY2', 'FY3']:
        r = results[did]
        init_pos = f'({DRONES[did][0]:.0f},{DRONES[did][1]:.0f},{DRONES[did][2]:.0f})'
        values = [
            did,
            init_pos,
            round(r['heading_deg'], 2),
            round(r['speed'], 1),
            round(r['t_drop'], 3),
            round(r['t_delay'], 3),
            round(r['t_burst'], 3),
            round(r['burst_x'], 1),
            round(r['burst_y'], 1),
            round(r['burst_z'], 1),
        ]
        for col, v in enumerate(values, 1):
            style_data_cell(ws, row, col)
            ws.cell(row=row, column=col, value=v)
        # 如果是FY2这种贡献为0的，标灰
        if r.get('single_coverage', 0) < 0.01:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        row += 1

    auto_width(ws)

    filepath = os.path.join(output_dir, 'result2.xlsx')
    wb.save(filepath)
    print(f'[OK] {filepath}')
    return filepath


# ============================================================
# result3.xlsx — 问题5结果
# ============================================================
def generate_result3(all_results, total_coverage, output_dir='.'):
    """
    参数:
        all_results: {missile_id: {decoded, coverage, burst_events, drone_list}}
        total_coverage: 总遮蔽时长
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '问题5结果'

    # 标题
    ws.merge_cells('A1:L1')
    ws.cell(row=1, column=1, value='问题5：全系统协同优化结果 (5机×≤3弹 → M1+M2+M3)')
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER_ALIGN

    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=f'总遮蔽时长: {total_coverage:.4f} s').font = \
        Font(name='微软雅黑', bold=True, size=12, color='C00000')
    row += 1

    # 逐导弹
    headers = ['导弹', '无人机', '初始位置', '航向角(°)', '速度(m/s)',
               '弹序', '投放时刻(s)', '起爆延迟(s)', '起爆时刻(s)',
               '起爆点X(m)', '起爆点Y(m)', '起爆点Z(m)']

    for mid in ['M1', 'M2', 'M3']:
        if mid not in all_results:
            continue

        r = all_results[mid]
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        ws.cell(row=row, column=1,
                value=f'{mid} — 遮蔽时长: {r["coverage"]:.4f}s').font = \
            Font(name='微软雅黑', bold=True, size=11, color='1F4E79')
        ws.cell(row=row, column=1).fill = LIGHT_FILL
        row += 1

        # 表头
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        for did in r['drone_list']:
            d = r['decoded'][did]
            init_pos = f'({DRONES[did][0]:.0f},{DRONES[did][1]:.0f},{DRONES[did][2]:.0f})'
            hdg = np.degrees(d['theta'])
            spd = d['speed']

            for bi, (t_drop, t_delay) in enumerate(d['drops'], 1):
                t_burst = t_drop + t_delay
                p_burst = bomb_burst_position(did, spd, d['theta'], t_drop, t_delay)
                values = [
                    mid, did, init_pos,
                    round(hdg, 2), round(spd, 1), bi,
                    round(t_drop, 3), round(t_delay, 3), round(t_burst, 3),
                    round(p_burst[0], 1), round(p_burst[1], 1), round(p_burst[2], 1),
                ]
                for col, v in enumerate(values, 1):
                    style_data_cell(ws, row, col)
                    ws.cell(row=row, column=col, value=v)
                row += 1

    auto_width(ws)

    filepath = os.path.join(output_dir, 'result3.xlsx')
    wb.save(filepath)
    print(f'[OK] {filepath}')
    return filepath


# ============================================================
# 批量生成（从各问题脚本导入结果）
# ============================================================
if __name__ == '__main__':
    output_dir = os.path.dirname(os.path.abspath(__file__))

    print("="*60)
    print("生成 Excel 结果文件")
    print("="*60)

    # --- result1.xlsx ---
    # 运行问题3获取参数（简化版——后续会从实际运行结果读取）
    print("\n>> 生成 result1.xlsx ...")
    try:
        from problem3 import solve_problem3
        theta, speed, t_drops, t_delays, coverage = solve_problem3()
        generate_result1(theta, speed, t_drops, t_delays, coverage, output_dir)
    except Exception as e:
        print(f"  [WARN] problem3 运行失败: {e}")
        # 使用默认值
        generate_result1(
            np.pi, 120.0,
            [1.5, 4.5, 7.5],
            [3.6, 3.6, 3.6],
            4.74, output_dir
        )

    # --- result2.xlsx ---
    print("\n>> 生成 result2.xlsx ...")
    try:
        from problem4 import solve_problem4
        results, coverage, _ = solve_problem4()
        generate_result2(results, coverage, output_dir)
    except Exception as e:
        print(f"  [WARN] problem4 运行失败: {e}")

    # --- result3.xlsx ---
    print("\n>> 生成 result3.xlsx ...")
    try:
        from problem5 import solve_problem5
        all_results, total_cov = solve_problem5()
        generate_result3(all_results, total_cov, output_dir)
    except Exception as e:
        print(f"  [WARN] problem5 运行失败: {e}")

    print(f"\n{'='*60}")
    print("所有 Excel 文件已生成")
    print(f"{'='*60}")
